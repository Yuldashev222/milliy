import openpyxl

book = openpyxl.Workbook()
sheet = book.active
for row in range(1, 10):
    step_col = 1
    for col in range(1, 10):
        sheet.cell(row, step_col).value = "{} * {}".format(col, row)
        sheet.cell(row, step_col + 1).value = "{}".format(col * row)
        step_col += 2
book.save("test.xlsx")
book.close()

# import openpyxl
#
# users = []
# users_book = openpyxl.open("test.xlsx")
# sheet = users_book.active
# user = input("enter name: ")
# users.append(user)
# print("added")
# sheet[1][0].value = user
# row = 2
# while True:
#     choice = input("continue or no?: ")
#     if choice == "no":
#         break
#     else:
#         user = input("enter name: ")
#         users.append(user)
#         sheet[row][0].value = user
#         print("added")
#         row += 1
#
# users_book.save("test.xlsx")
# users_book.close()
